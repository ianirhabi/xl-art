import argparse

import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from src.utils import rgb2hex


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--image',
        type=str,
        help='path to the input image',
    )
    parser.add_argument(
        '--reduce',
        type=float,
        default=1,
        help='to reduce size =>(h/reduce, w/reduce)',
    )
    parser.add_argument(
        '--sheet',
        type=str,
        default='art',
        help='sheet name',
    )
    parser.add_argument(
        '--output',
        type=str,
        default='output.xlsx',
        help='output path with file name',
    )
    args = parser.parse_args()

    img = cv2.imread(args.image)
    h, w = img.shape[:2]
    print(f'=========== Input image size = ({h}, {w}) ===========')
    h, w = int(h / args.reduce), int(w / args.reduce)  # 27 48
    print(f'=========== Output size = ({h}, {w}) ===========')
    img = cv2.resize(img, (w, h), interpolation=cv2.INTER_AREA)  # 48 27
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    wb = Workbook() 
    ws = wb.active
    ws.title = args.sheet

    for i, row in enumerate(ws[f'{get_column_letter(1)}1:{get_column_letter(w)}{str(h)}']):
        for j, cell in enumerate(row):
            print(f'[{i:5d},{j:5d}] Processing ...', end='\r')
            r, g, b = img[i, j]
            colorhex = rgb2hex(r,g,b)
            fill_gen = PatternFill(fill_type='solid', start_color=colorhex, end_color=colorhex)
            cell.fill = fill_gen

    for i in range(1, (w+1)):
        ws.column_dimensions[get_column_letter(i)].width = 1
        
    for i in range(1, (h+1)):
        ws.row_dimensions[i].height = 5

    wb.save(args.output)
    wb.close()
    print('=========== Finish ===========')

# import argparse
# import cv2
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter
# from src.utils import rgb2hex  # Assuming rgb2hex function is defined in src.utils module

# if __name__ == '__main__':
#     parser = argparse.ArgumentParser()
#     parser.add_argument('--image', type=str, help='path to the input image')
#     parser.add_argument('--sheet', type=str, default='art', help='sheet name')
#     parser.add_argument('--output', type=str, default='output.xlsx', help='output path with file name')
#     args = parser.parse_args()

#     # Load and resize the image
#     img = cv2.imread(args.image)
#     if img is None:
#         print(f"Error: Failed to load image from {args.image}")
#         exit(1)

#     h, w = img.shape[:2]
#     print(f'=========== Input image size = ({h}, {w}) ===========')

#     # Calculate new dimensions
#     target_height = 100
#     target_width = 200
#     resize_ratio = min(target_height / h, target_width / w)
#     h_new = int(h * resize_ratio)
#     w_new = int(w * resize_ratio)
#     print(f'=========== Resized image size = ({h_new}, {w_new}) ===========')

#     img_resized = cv2.resize(img, (w_new, h_new), interpolation=cv2.INTER_AREA)
#     img_resized = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = args.sheet

#     # Iterate through each pixel and fill Excel cells with corresponding color
#     for i, row in enumerate(ws[f'A1:{get_column_letter(w_new)}{h_new}']):
#         for j, cell in enumerate(row):
#             print(f'[{i:5d},{j:5d}] Processing ...', end='\r')
#             if i < h_new and j < w_new:
#                 r, g, b = img_resized[i, j]
#                 colorhex = rgb2hex(r, g, b)  # Assuming rgb2hex function converts RGB to hex color
#                 fill = PatternFill(fill_type='solid', start_color=colorhex, end_color=colorhex)
#                 cell.fill = fill

#     # Adjust column width and row height
#     for i in range(1, (w_new + 1)):
#         ws.column_dimensions[get_column_letter(i)].width = 1

#     for i in range(1, (h_new + 1)):
#         ws.row_dimensions[i].height = 5

#     # Save workbook and close
#     wb.save(args.output)
#     wb.close()

#     print('=========== Finish ===========')

# import argparse
# import cv2
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter
# from src.utils import rgb2hex  # Assuming rgb2hex function is defined in src.utils module

# def resize_image(img_path, target_height, target_width):
#     # Load and resize the image
#     img = cv2.imread(img_path)
#     if img is None:
#         print(f"Error: Failed to load image from {img_path}")
#         return None
    
#     h, w = img.shape[:2]
#     print(f'=========== Input image size = ({h}, {w}) ===========')

#     resize_ratio = min(target_height / h, target_width / w)
#     h_new = int(h * resize_ratio)
#     w_new = int(w * resize_ratio)
#     print(f'=========== Resized image size = ({h_new}, {w_new}) ===========')

#     img_resized = cv2.resize(img, (w_new, h_new), interpolation=cv2.INTER_LINEAR)  # Use INTER_LINEAR for better quality
#     img_resized = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)
    
#     return img_resized, h_new, w_new

# if __name__ == '__main__':
#     parser = argparse.ArgumentParser()
#     parser.add_argument('--image', type=str, help='path to the input image')
#     parser.add_argument('--sheet', type=str, default='art', help='sheet name')
#     parser.add_argument('--output', type=str, default='output.xlsx', help='output path with file name')
#     args = parser.parse_args()

#     # Resize the image
#     img_resized, h_new, w_new = resize_image(args.image, 300, 400)  # Target dimensions 300x400
#     if img_resized is None:
#         exit(1)

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = args.sheet

#     # Iterate through each pixel and fill Excel cells with corresponding color
#     for i, row in enumerate(ws.iter_rows(min_row=1, max_row=h_new, min_col=1, max_col=w_new)):
#         for j, cell in enumerate(row):
#             print(f'[{i:5d},{j:5d}] Processing ...', end='\r')
#             r, g, b = img_resized[i, j]
#             colorhex = rgb2hex(r, g, b)  # Assuming rgb2hex function converts RGB to hex color
#             fill = PatternFill(fill_type='solid', start_color=colorhex, end_color=colorhex)
#             cell.fill = fill

#     # Adjust column width and row height
#     for i in range(1, (w_new + 1)):
#         ws.column_dimensions[get_column_letter(i)].width = 2  # Adjust column width as needed

#     for i in range(1, (h_new + 1)):
#         ws.row_dimensions[i].height = 10  # Adjust row height as needed

#     # Save workbook and close
#     wb.save(args.output)
#     wb.close()

#     print(f'=========== Finish: Image resized and saved to {args.output} ===========')
